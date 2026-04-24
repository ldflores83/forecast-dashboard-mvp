"""
Salesforce Field Population Audit — Opportunity & Account
Auth: SID-based (refresh browser cookie every ~2hrs)
Output: sf_field_audit.xlsx with one tab per object + Summary tab

Sampling strategy: fetches the most recent N records by CreatedDate DESC
to avoid bias toward old records where custom fields were not yet in use.
"""

from simple_salesforce import Salesforce
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── AUTH ──────────────────────────────────────────────────────────────────────
SESSION_ID = "REMOVED_SF_SESSION_ID"  # F12 → Application → Cookies → copy "sid" value (starts with 00D)
sf = Salesforce(instance="qad.my.salesforce.com", session_id=SESSION_ID)

# ── CONFIG ────────────────────────────────────────────────────────────────────
SAMPLE_LIMIT  = 15000  # most recent records to sample for population rate
ID_BATCH_SIZE = 150    # max IDs per WHERE IN clause
FIELD_CHUNK   = 150    # max fields per SELECT (SOQL limit ~200, keeping buffer)
OUTPUT_FILE   = "sf_field_audit.xlsx"
OBJECTS       = ["Opportunity", "Account"]

# Key fields for the Revenue Intelligence dashboard — highlighted in output
DASHBOARD_FIELDS = {
    "Opportunity": [
        "StageName", "Amount", "Amount_converted__c", "CloseDate",
        "Primary_Opp_Value_Stream__c", "Opportunity_Owner_Business_Group__c",
        "ForecastCategoryName", "ForecastCategory",
        "Type", "LeadSource", "Probability",
        "Prior_Contract_End_Date__c", "ATR_Value__c", "ATR_Value_converted__c",
        "Total_Bookings_Net__c", "Substage__c",
        "Loss_Reason__c", "Cancellation_Reason__c",
        "Competitor__c", "IsClosed", "IsWon",
        "Fiscal_Period__c", "Fiscal_Year__c",
        "OwnerId", "AccountId", "Name",
    ],
    "Account": [
        "Name", "Type", "BillingCountry", "BillingCountryCode",
        "ERP_Customer_Base__c", "SC_Customer_Base__c",
        "Recurring_Rev_Customer_Base__c", "QAD_Customer_Base__c",
        "Global_HQ_18_ID__c", "Site_Type__c",
        "Status__c", "ARR__c", "Health_Score__c",
        "CS_Risk_Flag__c", "Health_Trend__c", "CS_Lifecycle_Stage__c",
        "Opportunity_Owner_Business_Group__c",
        "OwnerId", "CreatedDate",
    ]
}

# ── STYLES ────────────────────────────────────────────────────────────────────
def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF")
        )

def pct_fill(pct):
    if pct >= 80:  return "C6EFCE", "276221"
    if pct >= 40:  return "FFEB9C", "9C6500"
    return "FFC7CE", "9C0006"

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ── AUDIT FUNCTION ────────────────────────────────────────────────────────────
def audit_object(obj_name):
    print(f"\n── Auditing {obj_name} ──")

    # Fetch schema
    describe    = getattr(sf, obj_name).describe()
    fields      = describe["fields"]
    field_names = [f["name"] for f in fields]
    print(f"  Schema: {len(fields)} fields")

    # Step 1 — fetch most recent record IDs ordered by CreatedDate DESC
    print(f"  Fetching {SAMPLE_LIMIT:,} most recent IDs...")
    id_result  = sf.query_all(f"SELECT Id FROM {obj_name} ORDER BY CreatedDate DESC LIMIT {SAMPLE_LIMIT}")
    record_ids = [r["Id"] for r in id_result["records"]]
    total_org  = id_result["totalSize"]
    n          = len(record_ids)
    print(f"  Sample: {n:,} records (total in org: {total_org:,})")

    # Step 2 — batch IDs, fetch all fields in column chunks
    id_batches   = [record_ids[i:i+ID_BATCH_SIZE] for i in range(0, n, ID_BATCH_SIZE)]
    field_chunks = [field_names[i:i+FIELD_CHUNK]  for i in range(0, len(field_names), FIELD_CHUNK)]
    sample_data  = {f: [] for f in field_names}

    total_batches = len(id_batches)
    for b_idx, id_batch in enumerate(id_batches):
        if b_idx % 10 == 0:
            print(f"  Fetching batch {b_idx+1}/{total_batches}...")
        id_list = "', '".join(id_batch)
        for chunk in field_chunks:
            soql = f"SELECT {', '.join(chunk)} FROM {obj_name} WHERE Id IN ('{id_list}')"
            try:
                result = sf.query_all(soql)
                for record in result["records"]:
                    for field in chunk:
                        sample_data[field].append(record.get(field))
            except Exception as e:
                print(f"  Warning — batch {b_idx}, chunk error: {e}")
                for field in chunk:
                    sample_data[field].extend([None] * len(id_batch))

    # Build audit rows
    dashboard_set = set(DASHBOARD_FIELDS.get(obj_name, []))
    rows = []
    for f in fields:
        api       = f["name"]
        vals      = sample_data.get(api, [])
        populated = sum(1 for v in vals if v is not None and v != "" and v != [])
        pct       = round(populated / n * 100, 1) if n > 0 else 0

        picklist = ""
        if f.get("picklistValues"):
            active = [p["value"] for p in f["picklistValues"] if p.get("active")]
            picklist = " | ".join(active[:15])
            if len(active) > 15:
                picklist += f" (+{len(active)-15} more)"

        rows.append({
            "Dashboard Field": "★" if api in dashboard_set else "",
            "Field Label":     f.get("label", ""),
            "API Name":        api,
            "Type":            f.get("type", ""),
            "Is Custom":       "Yes" if f.get("custom") else "No",
            "Required":        "Yes" if not f.get("nillable") and not f.get("defaultedOnCreate") else "No",
            "Updateable":      "Yes" if f.get("updateable") else "No",
            "Description":     f.get("inlineHelpText") or "",
            "Picklist Values": picklist,
            f"% Populated (n={n})": pct,
        })

    rows.sort(key=lambda x: (-1 if x["Dashboard Field"] == "★" else 0, -x[f"% Populated (n={n})"]))
    return rows, n

# ── BUILD EXCEL ────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
summary_data = {}

for obj_name in OBJECTS:
    rows, n = audit_object(obj_name)
    summary_data[obj_name] = rows

    ws      = wb.create_sheet(title=obj_name)
    pct_col = f"% Populated (n={n})"
    headers = ["Dashboard Field", "Field Label", "API Name", "Type",
               "Is Custom", "Required", "Updateable", "Description",
               "Picklist Values", pct_col]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value     = f"{obj_name} Field Audit — {datetime.now().strftime('%Y-%m-%d')}  |  {n:,} most recent records (CreatedDate DESC)"
    c.font      = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    c.fill      = PatternFill("solid", start_color="0D1B2A")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, len(headers))
    ws.row_dimensions[2].height = 30

    # Data rows
    pct_idx = headers.index(pct_col) + 1
    for r_idx, row in enumerate(rows, 3):
        is_dashboard = row["Dashboard Field"] == "★"
        is_custom    = row["Is Custom"] == "Yes"
        pct_val      = row[pct_col]

        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            cell.font      = Font(size=9, name="Arial",
                                  bold=is_dashboard,
                                  color="1F3864" if is_dashboard else "000000")
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(h in ["Description", "Picklist Values"]))
            cell.border    = Border(bottom=Side(style="thin", color="E0E0E0"),
                                    right=Side(style="thin", color="E0E0E0"))
            if is_dashboard:
                cell.fill = PatternFill("solid", start_color="EBF3FB")
            elif is_custom:
                cell.fill = PatternFill("solid", start_color="FFFBF0")

        # Color-code the populated % cell
        bg, fg = pct_fill(pct_val)
        pct_cell = ws.cell(row=r_idx, column=pct_idx)
        pct_cell.value         = pct_val / 100
        pct_cell.number_format = "0.0%"
        pct_cell.fill          = PatternFill("solid", start_color=bg)
        pct_cell.font          = Font(size=9, name="Arial", bold=True, color=fg)
        pct_cell.alignment     = Alignment(horizontal="center", vertical="center")

    set_col_widths(ws, [6, 24, 32, 14, 8, 8, 10, 30, 50, 16])
    ws.freeze_panes    = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows)+2}"
    print(f"  Sheet '{obj_name}' written: {len(rows)} fields")

# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
ws = wb.create_sheet(title="Summary", index=0)

ws.merge_cells("A1:F1")
c = ws["A1"]
c.value     = f"Salesforce Field Audit — Summary  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Sample: {SAMPLE_LIMIT:,} most recent records"
c.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
c.fill      = PatternFill("solid", start_color="0D1B2A")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

headers_s = ["Object", "Total Fields", "Custom Fields", "Dashboard Fields",
             "Avg % Populated", "Fields > 80%"]
for col, h in enumerate(headers_s, 1):
    ws.cell(row=2, column=col, value=h)
style_header(ws, 2, len(headers_s))
ws.row_dimensions[2].height = 28

for r, (obj_name, rows) in enumerate(summary_data.items(), 3):
    pct_key   = [k for k in rows[0].keys() if k.startswith("% Populated")][0]
    total     = len(rows)
    custom    = sum(1 for x in rows if x["Is Custom"] == "Yes")
    dashboard = sum(1 for x in rows if x["Dashboard Field"] == "★")
    pcts      = [x[pct_key] for x in rows]
    avg_pct   = round(sum(pcts) / len(pcts), 1) if pcts else 0
    high_pct  = sum(1 for p in pcts if p >= 80)

    for col, val in enumerate([obj_name, total, custom, dashboard,
                                avg_pct / 100, high_pct], 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font      = Font(size=10, name="Arial", bold=(col == 1))
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                   vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color="E0E0E0"))
        if col == 5:
            cell.number_format = "0.0%"
            bg, fg = pct_fill(avg_pct)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.font = Font(size=10, name="Arial", bold=True, color=fg)

set_col_widths(ws, [18, 14, 14, 16, 18, 14])

# Legend section
ws["A6"] = "Legend"
ws["A6"].font = Font(bold=True, size=10, name="Arial")

legend = [
    ("★ Dashboard Field", "EBF3FB", "1F3864", "Key fields required for the Revenue Intelligence dashboard"),
    ("Custom Field",      "FFFBF0", "9C6500", "QAD custom fields"),
    (">= 80% populated",  "C6EFCE", "276221", "Well populated — reliable for reports and dashboards"),
    ("40-79% populated",  "FFEB9C", "9C6500", "Partially populated — validate before use"),
    ("< 40% populated",   "FFC7CE", "9C0006", "Low population rate — avoid using in dashboards"),
]
for i, (label, bg, fg, desc) in enumerate(legend, 7):
    cell_l           = ws.cell(row=i, column=1, value=label)
    cell_l.fill      = PatternFill("solid", start_color=bg)
    cell_l.font      = Font(size=9, name="Arial", bold=True, color=fg)
    cell_l.alignment = Alignment(vertical="center")
    ws.cell(row=i, column=2, value=desc).font = Font(size=9, name="Arial", color="444444")
    ws.merge_cells(f"B{i}:F{i}")

ws.row_dimensions[6].height = 20
ws.column_dimensions["B"].width = 60

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\nAudit saved -> {OUTPUT_FILE}")
print(f"Objects audited: {', '.join(OBJECTS)}")
print(f"Dashboard fields marked: {sum(len(DASHBOARD_FIELDS[o]) for o in OBJECTS)}")
print(f"Sample size: {SAMPLE_LIMIT:,} most recent records per object")